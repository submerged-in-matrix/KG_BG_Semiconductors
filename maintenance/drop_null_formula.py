"""
drop_null_formula.py — Remove rows with a null formula from all three files.

These are the 3 ids confirmed by trace_corruption.py as NaN in every stage
(mp-1009221, mp-1080032, mp-1179882) — genuine source gaps, not the Excel
date-conversion corruption (already fixed by fix_formula_corruption.py).

No hardcoded ids: discovers them fresh from whichever rows currently have
a null formula, in each file independently, so it stays correct if the
file changes.

Touches:
  - full_dataset_Bandgap_0_to_5.xlsx
  - full_dataset_Bandgap_0_to_5.csv
  - full_dataset_Bandgap_0_to_5_featurized.csv

Safe by default:
    python maintenance\\drop_null_formula.py            # dry run
    python maintenance\\drop_null_formula.py --commit   # writes (after .bak)
"""

import shutil
import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"

XLSX = DATA_DIR / "full_dataset_Bandgap_0_to_5.xlsx"
CLEAN_CSV = DATA_DIR / "full_dataset_Bandgap_0_to_5.csv"
FEAT_CSV = DATA_DIR / "full_dataset_Bandgap_0_to_5_featurized.csv"

ID_COL = "material_id"
FORMULA_COL = "formula"


def backup(path: Path):
    bak = path.with_suffix(path.suffix + ".bak")
    print(f"    backup {path.name} -> {bak.name}")
    shutil.copy2(path, bak)


def handle_csv(path: Path, commit: bool):
    print(f"\n{path.name}")
    df = pd.read_csv(path, dtype={ID_COL: str}, low_memory=False)
    if FORMULA_COL not in df.columns:
        print(f"    no '{FORMULA_COL}' column — skipping")
        return set()
    null_mask = df[FORMULA_COL].isna()
    ids = set(df.loc[null_mask, ID_COL])
    print(f"    rows total : {len(df):,}")
    print(f"    null formula: {int(null_mask.sum())}  -> {sorted(ids)}")
    if not ids:
        return ids
    if commit:
        backup(path)
        df = df.loc[~null_mask].reset_index(drop=True)
        df.to_csv(path, index=False)
        print(f"    written: {len(df):,} rows remain")
    return ids


def handle_xlsx(path: Path, commit: bool):
    print(f"\n{path.name}")
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("    openpyxl not installed — skipping (pip install openpyxl)")
        return set()

    print("    loading workbook (slow, ~150k rows)...")
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    try:
        id_i = header.index(ID_COL) + 1
        f_i = header.index(FORMULA_COL) + 1
    except ValueError:
        print(f"    header missing {ID_COL}/{FORMULA_COL}. Found: {header}")
        return set()

    rows_to_delete, ids = [], set()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=f_i).value
        if val is None:
            mid = ws.cell(row=r, column=id_i).value
            rows_to_delete.append(r)
            ids.add(mid)

    print(f"    rows total  : {ws.max_row - 1:,}")
    print(f"    null formula: {len(rows_to_delete)}  -> {sorted(ids)}")
    if not rows_to_delete:
        return ids

    if commit:
        backup(path)
        # delete bottom-up so earlier row indices stay valid
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r, 1)
        wb.save(path)
        print(f"    written: {ws.max_row - 1:,} rows remain")
    return ids


def main():
    commit = "--commit" in sys.argv
    mode = "COMMIT" if commit else "DRY RUN"
    print(f"=== drop_null_formula.py [{mode}] ===")

    for p in (XLSX, CLEAN_CSV, FEAT_CSV):
        if not p.exists():
            print(f"MISSING: {p}")
            return

    ids_csv = handle_csv(CLEAN_CSV, commit)
    ids_feat = handle_csv(FEAT_CSV, commit)
    ids_xlsx = handle_xlsx(XLSX, commit)

    print("\n" + "=" * 60)
    print("CROSS-CHECK — same ids dropped everywhere?")
    print("=" * 60)
    all_ids = ids_csv | ids_feat | ids_xlsx
    for mid in sorted(all_ids):
        where = []
        if mid in ids_csv: where.append("csv")
        if mid in ids_feat: where.append("featurized")
        if mid in ids_xlsx: where.append("xlsx")
        flag = "" if len(where) == 3 else "  <-- NOT IN ALL THREE, check manually"
        print(f"  {mid:<14} {where}{flag}")

    if commit:
        print("\nDone. Re-run trace_corruption.py to confirm 0 suspects remain.")
    else:
        print("\nDry run only. Re-run with --commit to apply.")


if __name__ == "__main__":
    main()
